import pandas as pd
from openpyxl import load_workbook

def verileri_birlestir(input_path, output_path):
    # Excel dosyasını oku
    df = pd.read_excel(input_path, header=None)
    
    # Tüm hücreleri tek sütunda topla ve benzersiz kayıtları al
    tum_veriler = df.stack().reset_index(drop=True).drop_duplicates()
    
    # Yeni Excel dosyası oluştur
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        tum_veriler.to_excel(writer, index=False, header=['Tüm Veriler'])
    
    print(f"Veriler {output_path} dosyasında AA sütununa kaydedildi!")

def aa_sutununa_yaz(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active
    
    # AA sütununu temizle
    for row in ws.iter_rows(min_col=27, max_col=27):
        for cell in row:
            cell.value = None
    
    # Verileri topla ve filtrele
    veriler = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and cell.row != 1 and cell.value not in veriler:
                veriler.append(cell.value)
    
    # AA sütununa yaz
    ws['AA1'] = 'Tüm Veriler'
    for idx, deger in enumerate(veriler, start=2):
        ws[f'AA{idx}'] = deger
    
    wb.save(output_path)
    print("İşlem tamamlandı!")

# Kullanım örneği
verileri_birlestir(
    input_path='2021REZERVASYON TAKİP.xlsx',
    output_path='Duzenlenmis_REZERVASYON.xlsx'
)

aa_sutununa_yaz('2021REZERVASYON TAKİP.xlsx', 'Duzenlenmis_REZERVASYON.xlsx')